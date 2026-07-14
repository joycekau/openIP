# -*- coding: utf-8 -*-
"""
365-day social media calendar generator for individual corax.live ecosystem products.

Usage:  python3 build_product_calendar.py <ProductKey>
        (run from a repo root; writes marketing/social-calendar/<file>.xlsx)

Product configs are grounded in the official marketing briefs + product repos:
  - CahtX  : spot DEX (perps COMING SOON) + no-code AI Bot Builder & copy-trading marketplace
  - Colony : high-frequency on-chain battle-royale betting game ("losing pays you back")
  - openIP : (set after inspecting the openIP repo)
"""
import sys, json
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

START_DATE = date(2026, 1, 1)

# ---------------------------------------------------------------- caption engine
ROLE_CAPTIONS = {
"SPOTLIGHT": [
 "Meet {feat}. 👀 {hook_cap} — on {product}. This is what {product} was built for. 🚀 {site}",
 "✨ Spotlight: {feat}. Imagine {hook} — on {product} that's just how it works. See it → {site}",
 "{feat} isn't a nice-to-have on {product}, it's the point. {hook_cap}. Try it → {site} 🔥",
 "New here? Start with {feat}. {hook_cap}. Welcome to {product}. 🌀 {site}",
 "This is {feat} on {product}: {hook}. Everything you need, one place. {site} 💠",
 "🔦 Feature drop → {feat}. {hook_cap}. And yes, it's live. {site}",
 "What if you could {hook_lower}? On {product} you can → {feat}. {site} ⚡",
 "Big things, one tap away: {feat}. {hook_cap}. Built for people, not middlemen. {product} · {site}",
 "Say hello to {feat}. 👋 {hook_cap}. The upgrade your feed's been waiting for. {site}",
 "{product} 101 → {feat}: {hook}. Simple to start, hard to leave. 😌 {site}",
],
"HOWTO": [
 "💡 Pro tip: use {feat} on {product} to {hook_lower}. Takes minutes. Here's how 👇 {site}",
 "How-To 🛠️ → getting the most from {feat}: {hook}. Bookmark this. {site}",
 "Did you know? On {product}, {feat} lets you {hook_lower} — no gatekeepers. 🙌 {site}",
 "Quick win 🎯 → try {feat}. Now you can {hook_lower}. That's the tip. ✅ {site}",
 "Power move 💪 → most people miss {feat}. {hook_cap}. Go try it on {product} → {site}",
 "Hidden gem 💎 → {feat}. {hook_cap}. Reply if this helped, we'll drop another. {site}",
 "Onboarding tip 🐣 → try {feat} first on {product}. Easiest 'wow': {hook}. {site}",
 "Learn it once, use it forever: {feat}. {hook_cap}. Save & share. 🔖 {site}",
 "3-step how-to 📲 → {feat} lets you {hook_lower}. Screenshot to remember. 📌 {site}",
 "Stop overcomplicating it. {feat} = {hook}. One flow on {product}. {site} ⚡",
],
"COMPARE": [
 "Tired of settling with {compare_cap}? {product} won't — thanks to {feat}: {hook}. Upgrade → {site} 🥇",
 "Still stuck with {compare_target}? 👀 {product} gives you {feat} — {hook} — without the baggage. {site}",
 "The honest comparison 🧾 {compare_target}: meh. {product}: {hook_lower}, and it's built for YOU. {site}",
 "{compare_cap} can't let you {hook_lower}. {product} can — it's called {feat}. 🔓 {site}",
 "{compare_cap} vs {product}? One feature says it all: {feat} ({hook}). 🎯 {site}",
 "Everyone's stuck on {compare_target}. Everything better is on {product} — starting with {feat}: {hook}. {site} 👥",
 "Why settle for {compare_target} when {product} gives you {feat}? {hook_cap}. Switch → {site}",
 "{compare_cap} was the old way. {product} is the upgrade: {feat}, so you {hook_lower}. 🚀 {site}",
 "Same thrill you wanted from {compare_target}, minus the catch, plus {feat}. {hook_cap}. 😏 {site}",
 "A fair fight ⚔️ {compare_target} vs {product}. {feat} wins it: {hook}. {site}",
],
"DEEPDIVE": [
 "🧠 Under the hood: {feat}. {hook_cap}. We engineered this so it's powerful AND yours. {site}",
 "Let's talk tech 🔬 → {feat} on {product} means {hook}. Real infra, not marketing. {site}",
 "Why {feat} matters: {hook}. Defaults decide everything — {product} defaults to you. 🛡️ {site}",
 "Deep-dive 🌊 → how {feat} actually works: {hook}. No black boxes. {site}",
 "Nerd corner 🤓 → {feat}. {hook_cap}. This should be the standard, not the exception. {site}",
 "For the pros 👷 → {feat} on {product}: {hook}. Web3-native, actually usable. {site}",
 "Transparency > hype 🔎 → the 'how' behind {feat}: {hook}. You deserve to know. {site}",
 "One tweetable truth: {feat} = {hook}. The tech is deep; the promise is simple. 🤝 {site}",
 "Trust note 🔒 → {feat} gives you {hook}. Your keys, your rules. {site}",
 "{eco} 🧩 {feat} is how it clicks: {hook}. {site}",
],
"STORY": [
 "Real scenario 🎬 → {hook_cap}. That moment usually starts with {feat} on {product}. {site}",
 "Meet a {product} user who used {feat} to {hook_lower}. Game changed. 💼 {site}",
 "A day on {product} ☀️ → {feat} kicks in and suddenly you {hook_lower}. That's the magic. {site}",
 "From the old grind to one clean flow. The turning point? {feat}: {hook}. 🙌 {site}",
 "True story: someone came for one thing and stayed for {feat} — {hook}. Your turn? 👀 {site}",
 "The 'aha' moment 💡 usually starts with {feat}: {hook}. Find yours → {site}",
 "Traders, creators, players, builders — all using {feat} to {hook_lower}. One {product}. 🌍 {site}",
 "Community hero moment 🦸 → they showed off {feat} and everyone went 'wait, it does THAT?' Yep: {hook}. {site}",
 "Picture this 🌆 → {hook}. {feat} makes it real on {product}. {site}",
 "Use-case of the week 📖 → {feat}: {hook}. Steal this. {site}",
],
"ENGAGE": [
 "POLL 🗳️ → what matters most: A) {hook_lower} B) low fees C) real ownership D) all of it? ({product} says D 😏) {site}",
 "Quick question 🤔 → if you had {feat} ({hook}), what would you do first? Reply 👇 {site}",
 "Fill in the blank ✍️ → 'I'd go all-in on {product} if it had ______.' (It probably does.) {site}",
 "This or that 🔀 → the old way, OR {feat} ({hook})? Vote ❤️ or 🔁. {site}",
 "Tag someone 👀 who doesn't believe you can {hook_lower}. Then show them {feat}. {site}",
 "GM ☀️ fam! What should we spotlight next on {product}: {feat} or something else? Comment 🗨️ {site}",
 "Hot take 🌶️ → {hook}. Agree? {feat} on {product} makes it real. {site}",
 "Riddle 🧩 → what lets you {hook_lower} in one move? (It's {feat} on {product} 😉) Answer below. {site}",
 "Retweet 🔁 if you're done settling. Reply with the ONE feature that'd make you switch to {product}. 👂 {site}",
 "Be honest 😅 → how are you doing this today? {product} does it better. Drop your setup 👇 {site}",
],
"CTA": [
 "Ready? 🚀 Start with {product} at {site} — then hang with us on {platforms} (links in bio). 👥",
 "Your invite is open 💌 → {site}. {hook_cap} with {feat}. Say hi on {platforms}. 🫶",
 "Don't just watch — jump in. 🔓 {site}. Follow on {platforms} for drops & community. 🐦💬",
 "Early is a superpower. ⏳ Get into {product} now ({site}) and grow with us. {platforms} pinned. 📌",
 "One tap to something better → {site}. Then plug into the movement on {platforms}. 🪑",
 "Join the people going all-in. 🌐 {site}. New here? {platforms} welcomes you. ✨",
 "Weekend nudge 🧘 → start in seconds at {site}. Meet us on {platforms}. 🔁",
 "The app is the start. The community is the point. Begin at {site}, belong on {platforms}. 💛",
 "Read this far? You're curious 😄 feed it → {site}. Then GM us on {platforms}. 🌀",
 "Play/trade/build on YOUR terms. {site} to begin · {platforms} to belong. Welcome to {product}. 🔥",
],
}
ROLE_IMG = {
"SPOTLIGHT":"Hero product shot: {subject}. Bold headline '{feat}', small tag '{site}'. {style}",
"HOWTO":"Instructional carousel cover: {subject}, with a numbered 1-2-3 step overlay and a lightbulb icon. Text '{feat} in 3 steps'. {style}",
"COMPARE":"Split-screen VS graphic: left a dull {compare_target}, right a glowing {product} screen featuring {subject}. 'VS' badge center, text 'Upgrade → {site}'. {style}",
"DEEPDIVE":"Techy explainer/blueprint illustration: {subject}, surrounded by circuit lines, tokens and data particles. Small label '{feat}'. {style}",
"STORY":"Lifestyle scene: a real, diverse person delighted while using a device showing {subject}. Warm relatable setting. Subtle {product} branding. {style}",
"ENGAGE":"Bold social poll graphic: big question text, 2-4 tappable option chips, emojis, clean negative space. Include {subject} as a small motif. {style}",
"CTA":"Punchy call-to-action poster: {subject} as hero background, giant 'Join {site}' button, small X (Twitter) bird + Telegram paper-plane icons. {style}",
}
WEEK_ROLES = ["SPOTLIGHT","HOWTO","COMPARE","DEEPDIVE","STORY","ENGAGE","CTA"]
CORE = ["#Web3","#Crypto","#Telegram","#OnX","#Community"]

def cap_first(s): return s[:1].upper()+s[1:] if s else s
def build_tags(cfg, role, i):
    t=[cfg["brand_tag"]]; pt=cfg["tags"]
    t+=[pt[i%len(pt)], pt[(i+3)%len(pt)], CORE[i%len(CORE)], CORE[(i+2)%len(CORE)]]
    t.append({"SPOTLIGHT":"#Web3","HOWTO":"#HowTo","COMPARE":"#SwitchNow","DEEPDIVE":"#OnChain",
              "STORY":"#RealStories","ENGAGE":"#YourTurn","CTA":"#JoinNow"}[role])
    seen,out=set(),[]
    for x in t:
        if x.lower() not in seen: seen.add(x.lower()); out.append(x)
    return " ".join(out[:8])

WRAP=Alignment(wrap_text=True,vertical="top")
CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True)
_th=Side(style="thin",color="D9D9D9"); BORDER=Border(left=_th,right=_th,top=_th,bottom=_th)
def tint(h,f):
    r,g,b=int(h[0:2],16),int(h[2:4],16),int(h[4:6],16)
    return f"{int(r+(255-r)*f):02X}{int(g+(255-g)*f):02X}{int(b+(255-b)*f):02X}"

def build(cfg):
    feats=list(cfg["features"].keys()); rows=[]
    for i in range(365):
        d=START_DATE+timedelta(days=i); role=WEEK_ROLES[d.weekday()]
        feat=feats[i%len(feats)]; hook,subject=cfg["features"][feat]
        theme=cfg["themes"][d.month-1][0]
        tmpl=ROLE_CAPTIONS[role][(i//7)%len(ROLE_CAPTIONS[role])]
        cap=tmpl.format(product=cfg["name"],feat=feat,hook=hook,hook_cap=cap_first(hook),
            hook_lower=hook[0].lower()+hook[1:],site=cfg["site"],compare_target=cfg["compare_target"],
            compare_cap=cap_first(cfg["compare_target"]),eco=cfg["ecosystem"].split(".")[0]+".",platforms=cfg["platforms"])
        img=ROLE_IMG[role].format(feat=feat,subject=subject,site=cfg["site"],product=cfg["name"],
            compare_target=cfg["compare_target"],style=cfg["style"])
        rows.append({"Day":i+1,"Date":d.strftime("%Y-%m-%d"),"Weekday":d.strftime("%a"),
            "Month Theme":theme,"Content Type":cfg["labels"][role],"Feature Focus":feat,
            "Platform":["X (primary) + Telegram","Telegram (primary) + X","X + Telegram","X + Telegram + WhatsApp/WeChat status"][i%4],
            "Best Time":["9:00 AM","12:30 PM","6:00 PM","8:00 PM"][i%4],"Caption":cap,
            "Hashtags":build_tags(cfg,role,i),"Image Prompt (Grok / Gemini / Cora AI)":img,
            "CTA":f"Start at {cfg['site']} · Follow on X · Join Telegram"})
    wb=Workbook(); ws=wb.active; ws.title="365-Day Calendar"
    hdr=list(rows[0].keys()); ws.append(hdr)
    HF=PatternFill("solid",fgColor=cfg["brand_color"]); HFONT=Font(bold=True,color=cfg["accent"],size=11)
    for c,_ in enumerate(hdr,1):
        x=ws.cell(row=1,column=c); x.fill=HF; x.font=HFONT; x.alignment=CENTER; x.border=BORDER
    ba,bb=tint(cfg["accent"],.86),tint(cfg["accent"],.74)
    for ri,row in enumerate(rows,2):
        fill=PatternFill("solid",fgColor=(ba if date.fromisoformat(row["Date"]).month%2==0 else bb))
        for ci,k in enumerate(hdr,1):
            x=ws.cell(row=ri,column=ci,value=row[k]); x.alignment=WRAP; x.border=BORDER; x.fill=fill
    W={"Day":6,"Date":12,"Weekday":9,"Month Theme":22,"Content Type":26,"Feature Focus":26,"Platform":26,
       "Best Time":10,"Caption":74,"Hashtags":42,"Image Prompt (Grok / Gemini / Cora AI)":82,"CTA":34}
    for ci,k in enumerate(hdr,1): ws.column_dimensions[get_column_letter(ci)].width=W.get(k,18)
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(hdr))}1"
    # Read Me
    s=wb.create_sheet("Read Me — Product & Strategy"); s.column_dimensions["A"].width=28; s.column_dimensions["B"].width=112
    r=[1]
    def kv(a,b,bold=True):
        s.cell(row=r[0],column=1,value=a).font=Font(bold=bold,color=cfg["brand_color"],size=11)
        s.cell(row=r[0],column=1).alignment=WRAP; s.cell(row=r[0],column=2,value=b).alignment=WRAP
        s.row_dimensions[r[0]].height=54; r[0]+=1
    s.cell(row=r[0],column=1,value=f"{cfg['name']} · 365-Day Social Media Plan").font=Font(bold=True,size=20,color=cfg["brand_color"]); r[0]+=1
    s.cell(row=r[0],column=1,value=cfg["tagline"]).font=Font(italic=True,size=11,color="444444"); r[0]+=2
    kv("What it is",cfg["positioning"])
    if cfg.get("botbuilder"): kv("Flagship: Bot Builder",cfg["botbuilder"])
    kv("Target Segments",cfg["audience"])
    kv("Ecosystem Fit (→ corax.live)",cfg["ecosystem"])
    kv("Objective",f"Grow {cfg['platforms']} followers & engagement and drive sign-ups to {cfg['site']}, feeding the wider corax.live ecosystem.")
    kv("Weekly Rhythm"," · ".join(f"{dd}: {cfg['labels'][WEEK_ROLES[i]]}" for i,dd in enumerate(['Mon','Tue','Wed','Thu','Fri','Sat','Sun'])))
    kv("Monthly Themes"," · ".join(f"M{i+1}: {n}" for i,(n,_) in enumerate(cfg["themes"])))
    kv("How to use","1 row = 1 day / 1 post. Copy Caption + Hashtags into X/Telegram; paste the Image Prompt into Grok, Gemini or Cora AI; post at the suggested time (adjust to your timezone). Repurpose winners as X threads & WhatsApp/WeChat statuses. Reply in the first hour.")
    if cfg.get("compliance"): kv("⚠️ Compliance / accuracy",cfg["compliance"])
    kv("⚠️ Confirm before posting",cfg["site_note"])
    # Hashtag bank
    h=wb.create_sheet("Hashtag & Asset Bank"); h.column_dimensions["A"].width=30; h.column_dimensions["B"].width=95
    h.cell(row=1,column=1,value="Reusable Hashtag Sets").font=Font(bold=True,size=14,color=cfg["brand_color"])
    bank=[("Always-on (brand)",cfg["brand_tag"]+" "+" ".join(cfg["tags"][:3])),
          ("Ecosystem","#CoraX #coraxlive #Web4 #Web3 #Crypto"),
          ("Community/CTA","#Telegram #OnX #JoinNow #GM #Web3Community"),
          ("Product-specific"," ".join(cfg["tags"]))]
    rr=2
    for k,v in bank:
        h.cell(row=rr,column=1,value=k).font=Font(bold=True,color=cfg["brand_color"]); h.cell(row=rr,column=1).alignment=WRAP
        h.cell(row=rr,column=2,value=v).alignment=WRAP; rr+=1
    rr+=1
    h.cell(row=rr,column=1,value="Brand image style (append to prompts)").font=Font(bold=True,color=cfg["brand_color"])
    h.cell(row=rr,column=1).alignment=WRAP; h.cell(row=rr,column=2,value=cfg["style"]).alignment=WRAP
    wb.save(cfg["file"]); return cfg["file"], len(rows)

# ---------------------------------------------------------------- product configs
PRODUCTS = {}

PRODUCTS["CahtX"] = {
 "name":"CahtX","file":"marketing/social-calendar/cahtx_365_social_calendar.xlsx",
 "site":"cahtx.app","site_note":"Confirm the real CahtX URL/handle (placeholder 'cahtx.app'); one find-replace updates all rows.",
 "platforms":"X & Telegram","brand_tag":"#CahtX",
 "tags":["#BotBuilder","#CopyTrading","#TradingBots","#NoCode","#SpotDEX","#AITrading","#DeFi","#Backtesting"],
 "brand_color":"0E2A2A","accent":"38F5C9",
 "tagline":"A spot DEX + no-code AI trading bots you can build, run, copy & monetize.",
 "positioning":("CahtX is a non-custodial SPOT DEX with a no-code AI Bot Builder. Describe a strategy in plain English and Cora turns it into a live bot that trades 24/7 on a managed engine — backtest it first, add built-in risk controls, or copy top creators' bots from the marketplace. (Perpetual futures are COMING SOON.)"),
 "botbuilder":("Describe your strategy in plain English → Cora builds the bot → it trades for you 24/7. Backtest before risking a cent, run it on a managed engine with hard risk controls, and publish winning bots to the marketplace to earn recurring income (creators keep 70%). Copy traders follow proven bots in one click."),
 "audience":("Everyday traders (build a bot with zero code — 'trade like a quant without being one'); Strategy creators (publish to the marketplace, earn recurring monthly income, keep 70%); Copy traders (one-click auto-copy top bots); Advanced builders (custom rule-based logic, full control, zero infra). Plus spot traders who want self-custody."),
 "ecosystem":("CahtX is the FINANCIAL + AUTOMATION engine of the corax.live ecosystem. It's non-custodial (your keys, your funds); Cora AI powers the plain-English bot builder shared across CoraX; bots and copy-trading can plug into your CoraX communities, Telegram & X; earnings compound across the ecosystem."),
 "compare_target":"a centralized exchange",
 "labels":{"SPOTLIGHT":"Feature Spotlight","HOWTO":"Bot-Builder How-To","COMPARE":"CahtX vs a CEX",
           "DEEPDIVE":"Tech Deep-Dive","STORY":"Trader / Creator Story","ENGAGE":"Engagement / Poll","CTA":"CTA / Join Us"},
 "compliance":"Not financial advice. Never guarantee profits or quote fixed returns; trading is risky. 'Perpetual futures' and 'PRO technical indicators' are COMING SOON — label them as roadmap, not live. Backtests are historical, not a promise of future results.",
 "style":("Brand look: sleek dark trading-terminal UI, deep teal-black background, neon mint-green & electric-cyan "
          "candlestick glows, subtle grid, glassmorphism cards, a friendly robot/bot motif, clean mono+sans type, "
          "premium quant-meets-chat aesthetic, 8k, crisp, high contrast, minimal, professional."),
 "themes":[
  ("Meet CahtX","Spot DEX + no-code trading bots — the core promise."),
  ("No-Code Bot Builder","Plain-English strategy → live bot."),
  ("Strategy Templates","Market-making, trend, grid, DCA out of the box."),
  ("Backtest & AI Coaching","Test on real history, get AI feedback."),
  ("Bots That Run 24/7","The managed engine & live dashboard."),
  ("Copy-Trading Marketplace","Follow & copy proven bots in one click."),
  ("Creators Earn 70%","Publish a bot, earn recurring income."),
  ("Safe by Design","Built-in risk controls & self-custody."),
  ("Spot Trading","Non-custodial spot swaps done right."),
  ("Advanced Builder Season","Custom rule-based logic for pros."),
  ("Creator Spotlight","Trader & creator success stories."),
  ("Year of CahtX","Milestones, roadmap & Perps coming soon."),
 ],
 "features":{
  "Plain-English Bot Builder":("describe a trading strategy in plain English and get a live bot — no code",
    "a trader typing a plain sentence that morphs into a glowing automated trading-bot flowchart"),
  "Strategy Templates":("start from Market-Making, Trend, Grid or DCA templates in a tap",
    "four glowing strategy template cards labelled Market Making, Trend, Grid and DCA"),
  "Backtesting & AI Coaching":("backtest on real market history and get plain-language AI coaching before going live",
    "an equity-curve chart with win/loss stats and an AI coach speech bubble giving feedback"),
  "24/7 Managed Bot Engine":("run your bot around the clock on a managed engine — nothing to host",
    "a friendly robot tending glowing trades around a 24/7 clock, server nodes humming behind"),
  "Copy-Trading Marketplace":("copy a top-performing bot in one click and let it trade for you",
    "a marketplace of ranked bot cards showing performance, with a one-tap 'Copy' button glowing"),
  "Creator Earnings (70%)":("publish your bot and keep 70% of monthly subscriber revenue",
    "a creator earnings dashboard with rising subscriber count and a recurring monthly-revenue chart"),
  "Built-In Risk Controls":("cap capital, set account stop-loss and auto-block tiny orders on every bot",
    "a trading bot inside a glowing safety shield with capital-cap and stop-loss dials"),
  "Custom Strategy Logic":("build advanced scaled entries, exits and position sizing with priority rules",
    "an advanced rule-builder canvas with layered order logic blocks snapping together"),
  "Non-Custodial Spot Trading":("trade spot straight from your own wallet — the platform never holds your funds",
    "a self-custody spot swap panel with a key icon, funds staying safely in the user's wallet"),
  "Perps (Coming Soon)":("perpetual futures are coming soon to CahtX",
    "a sleek 'Perps — Coming Soon' teaser banner over a futures chart with a countdown"),
 },
}

PRODUCTS["Colony"] = {
 "name":"Colony","file":"marketing/social-calendar/colony_365_social_calendar.xlsx",
 "site":"colony.land","site_note":"Domain 'colony.land' confirmed.",
 "platforms":"X & Telegram","brand_tag":"#Colony",
 "tags":["#ColonyLand","#Web3Gaming","#ProvablyFair","#BattleRoyale","#GameFi","#PlayOnChain","#COLONY","#Crypto"],
 "brand_color":"1A0B0B","accent":"58F542",
 "tagline":"High-frequency Web3 battle royale — bet the safe room, and losing pays you back.",
 "positioning":("Colony is a high-frequency on-chain battle royale. Every 45 seconds, 9 rooms appear — 8 get invaded by zombies, 1 is safe. Pick your room in the live betting window; winners split the pot. Lose? The Mining Pool automatically pays you back over time in COLONY. Provably fair on-chain randomness, non-custodial."),
 "botbuilder":"",
 "audience":("Web3 gamers & degens who want fast, low-friction on-chain action (a full round every 45s); players burned by riggable games who need provable fairness; casual players (the $1 tier) through to whales (the $100 tier), each in their own pool; and 'set-and-forget' grinders who love Auto Mode."),
 "ecosystem":("Colony is the GAMING + ENTERTAINMENT arm of the corax.live ecosystem. It's non-custodial and provably fair; players sign in and squad up through CoraX communities; the COLONY token economy (fixed supply, buyback-and-burn) and the Node Shareholder Program tie into the wider ecosystem; identity travels via openIP."),
 "compare_target":"a rigged casino game",
 "labels":{"SPOTLIGHT":"Gameplay Spotlight","HOWTO":"How to Play","COMPARE":"Colony vs Rigged Games",
           "DEEPDIVE":"Provably-Fair Tech","STORY":"Player Story","ENGAGE":"Engagement / Meme / Poll","CTA":"CTA / Play Now"},
 "compliance":"This is real-money on-chain betting — keep it responsible. Never promise winnings or 'guaranteed' recovery amounts/timing; Mining Pool recovery is gradual and paid in COLONY. Exclude wallet addresses, fee recipients, keeper/operator mechanics and uncalibrated volume figures. Add play-responsibly framing; respect age/geo rules.",
 "style":("Brand look: cinematic post-apocalyptic arcade art, dark grim ruins, toxic radioactive-green glow, "
          "neon HUD overlays, 9 doors/rooms motif, countdown timers, dramatic rim light, gritty textures, epic "
          "battle-royale energy, stylized 3D, 8k, high detail, high contrast, dynamic composition."),
 "themes":[
  ("Enter the Colony","Meet the 45-second battle royale."),
  ("How a Round Works","9 rooms, 1 safe, one decision."),
  ("Losing Pays You Back","The Mining Pool — the killer hook."),
  ("Play Your Level","Bet tiers from $1 to $100."),
  ("The Zombie King","The jackpot layer & hype moments."),
  ("Hands-Free: Auto Mode","Set strategy + limits, let it run."),
  ("We Can't Rig It","Provably fair & non-custodial."),
  ("The COLONY Token","Fixed supply, no tax, buyback & burn."),
  ("Own a Node","The Node Shareholder Program."),
  ("Pro-Player Season","Strategy, contrarian plays, tactics."),
  ("Player Spotlight","Wins, clips & community legends."),
  ("Year of the Colony","Milestones, roadmap & big drops."),
 ],
 "features":{
  "45-Second Rounds":("play a full round every 45 seconds — no lobbies, no waiting, pure adrenaline",
    "a glowing countdown timer over 9 doors as a crowd piles in and zombies loom in green fog"),
  "Pick the Safe Room":("make one decision — which of the 9 rooms is safe — and win the pot",
    "nine doors, eight leaking green zombie glow and one golden safe door, a hand reaching to choose"),
  "Losing Pays You Back":("lose a round and recover your stake over time through the Mining Pool in COLONY",
    "a red 'loss' turning into a glowing seed that grows COLONY coins inside a mining pool"),
  "Bet Tiers $1–$100":("play your level — $1 to $100, each tier in its own fair pool",
    "five stake-tier chips ($1 $5 $10 $50 $100), each feeding its own separate glowing prize pool"),
  "Zombie King Jackpot":("catch the Zombie King storming your room and win a massive jackpot",
    "a giant crowned Zombie King bursting into a room, raining a huge jackpot of coins on the players"),
  "Auto Mode":("set your room strategy, round limit and loss cap and play completely hands-free",
    "an autopilot control panel with room-strategy, max-rounds and max-loss dials, a safety toggle"),
  "Provably Fair":("trust the result — the safe room is on-chain random and nobody can rig it, not even us",
    "a glowing on-chain RNG dice sealed in a tamper-proof shield with the words 'we can't rig it'"),
  "Non-Custodial Vault":("your funds stay yours — deposit to a vault, play, and withdraw anytime on-chain",
    "a personal glowing vault whose key is held by the player, funds never leaving their control"),
  "COLONY Token":("clean tokenomics — fixed supply, no tax, with buyback-and-burn every round",
    "a COLONY coin above a supply curve that only points down, with a burn flame consuming tokens"),
  "Node Shareholder Program":("own a node and share in the growth of the game",
    "a premium glowing node badge beside a revenue-share pie, a community standing proudly behind it"),
 },
}

PRODUCTS["OneIP"] = {
 "name":"OneIP","file":"marketing/social-calendar/oneip_365_social_calendar.xlsx",
 "site":"oneip.io","site_note":("Domain 'oneip.io' confirmed. NOTE: some on-chain features run on Solana DEVNET "
   "(mainnet pending) — verify what's actually live before claiming it in a post."),
 "platforms":"X & Telegram","brand_tag":"#OneIP",
 "tags":["#Solana","#CreatorCoin","#KOL","#Launchpad","#TokenizeYourInfluence","#SOL","#CreatorEconomy","#Web3"],
 "brand_color":"1A1608","accent":"E9C46A",
 "tagline":"Tokenize your influence — launch your creator coin on Solana, with a 20% floor.",
 "positioning":("OneIP.io is a Solana creator-IP / KOL token launchpad + gmgn-style trading terminal + creator social platform. Creators launch their own token in one click — every coin ships with a 20% FLOOR treasury (it can't go to zero), verified socials that display across Phantom, Solflare, Backpack, Jupiter & Solscan, and a built-in creator feed. Non-custodial (Phantom), mobile, 7 languages, built for SE-Asia creator networks."),
 "botbuilder":"",
 "audience":("Creators / KOLs / influencers (esp. SE-Asia / WebTVAsia) who want to monetize their influence — launch a token, run a creator feed, and even auto-reply with a Pro AI clone; traders & degens who trade creator coins and track smart-money KOL wallets on the gmgn-style terminal; fans/communities who buy, hold and can redeem the on-chain floor."),
 "ecosystem":("OneIP is the CREATOR-TOKENIZATION / KOL arm of the broader corax.live ecosystem — creators turn audience & influence into an on-chain asset. It's non-custodial; the Pro AI-clone auto-reply is Cora/Claude-powered; creator communities can live alongside CoraX, and creator coins are tradable in the wider Web3 economy."),
 "compare_target":"a typical meme-coin launchpad",
 "labels":{"SPOTLIGHT":"Feature Spotlight","HOWTO":"Launch How-To","COMPARE":"OneIP vs Meme Launchpads",
           "DEEPDIVE":"On-Chain / Trust Tech","STORY":"Creator Story","ENGAGE":"Engagement / Poll","CTA":"CTA / Launch Now"},
 "compliance":("Not financial advice; token trading is risky. The 20% floor is a treasury backstop that reduces downside — NOT a guarantee of price or returns. Never promise gains. Some on-chain features are on Solana devnet / undeployed — don't claim mainnet features that aren't live. Never post API keys/secrets."),
 "style":("Brand look: premium dark-and-gold 'noble' fintech UI, near-black background with rich gold accents, "
          "Solana purple-green hints, gmgn-style data-terminal charts, verified/checkmark and floor motifs, "
          "glassmorphism cards, elegant sans type, trustworthy high-end aesthetic, 8k, crisp, high contrast."),
 "themes":[
  ("Meet OneIP","Tokenize your influence — the core promise."),
  ("One-Click Launch","Launch a creator coin in minutes."),
  ("The 20% Floor","Why your coin can't go to zero."),
  ("Verified Everywhere","Logo + socials across every wallet."),
  ("The Trading Terminal","gmgn-style trending & charts."),
  ("Follow the Smart Money","KOL wallet tracker & PnL."),
  ("Your Creator Feed","Posts, followers & subscriptions."),
  ("Your AI Clone","Pro AI auto-reply in your voice."),
  ("Safe & Verified","Floor redemption, RugCheck, attestation."),
  ("Non-Custodial & Global","Phantom, 7 languages, mobile."),
  ("Creator Spotlight","KOL & creator launch stories."),
  ("Year of OneIP","Milestones, roadmap & mainnet."),
 ],
 "features":{
  "One-Click Token Launch":("launch your own creator coin in one click — logo, site and socials attached",
    "a creator tapping a glowing gold Launch button as their coin mints with logo and social links attached"),
  "20% Floor Treasury":("every coin ships with a 20% floor treasury so it can't go to zero",
    "a gold coin resting on a glowing floor labelled '20%', a falling price arrow stopped dead by the floor"),
  "Cross-Wallet Credibility":("your logo and verified socials show in Phantom, Solflare, Backpack, Jupiter and Solscan",
    "one creator coin's logo and a verified checkmark appearing consistently across several wallet app screens"),
  "gmgn-Style Trading Terminal":("track trending coins and trade on a fast, pro-grade data terminal",
    "a sleek dark-and-gold trading terminal showing trending Solana tokens with live charts and stats"),
  "KOL Smart-Money Tracker":("see what top KOL wallets are buying, with a live PnL engine",
    "a smart-money leaderboard of KOL wallets showing live PnL and their latest buys and sells"),
  "Creator Social Feed":("post, grow followers and sell subscriptions on your own creator feed",
    "a creator social feed with posts, follow and subscribe buttons and an engaged fan community"),
  "AI Clone Auto-Reply":("a Pro AI clone that auto-replies to your community in your own voice",
    "a glowing AI twin of a creator automatically answering a stream of fan messages"),
  "Redeem the Floor":("holders can redeem the on-chain floor value anytime — real backing, not vibes",
    "a holder pressing a 'redeem floor' button and receiving backed value from a glowing on-chain vault"),
  "Security & Verification":("RugCheck badges and signed social attestation build trust at a glance",
    "a coin card showing a green RugCheck security shield and a verified-social checkmark"),
  "Non-Custodial · 7 Languages":("connect Phantom, keep full custody, and use it in 7 languages",
    "a Phantom wallet connect screen with a language selector showing EN 中文 ไทย Bahasa Tiếng-Việt 日本語 한국어"),
 },
}

if __name__ == "__main__":
    key = sys.argv[1] if len(sys.argv) > 1 else None
    if key == "ALL":
        for k,c in PRODUCTS.items(): print("Wrote", *build(c), f"[{k}]")
    elif key in PRODUCTS:
        print("Wrote", *build(PRODUCTS[key]), f"[{key}]")
    else:
        print("Usage: build_product_calendar.py <ProductKey>  |  keys:", ", ".join(PRODUCTS)); sys.exit(1)
